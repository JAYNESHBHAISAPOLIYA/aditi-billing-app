"""
Government RA Bill Excel Generator using openpyxl
Generates a multi-sheet Excel file in GUDCL Halol WSS project format.

Usage:
    from generate_ra_bill import generate_ra_bill
    excel_bytes = generate_ra_bill(bill_data)

    Or run standalone:
    python generate_ra_bill.py
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
AMOUNT_FORMAT = "##,##,##0.00"


def _merge(ws, row, col_start, col_end, value, bold=True, size=11,
           fill=None, align="center", height=None, wrap=False):
    """Merge cells, set value, style, and optionally row height."""
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    if col_end > col_start:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row, end_column=col_end
        )
    if height:
        ws.row_dimensions[row].height = height


def _border_cell(cell, value=None, bold=False, number_format=None,
                 align="left", fill=None):
    if value is not None:
        cell.value = value
    cell.border = THIN_BORDER
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill


def generate_ra_bill(bill_data: dict) -> bytes:
    """
    Generate a Government RA Bill Excel file.

    Args:
        bill_data (dict): {
            "bill_no": "1",
            "bill_period_from": "2024-04-01",
            "bill_period_to": "2024-09-30",
            "bill_date": "2024-10-01",
            "company_name": "GUJARAT URBAN DEVELOPMENT COMPANY, GANDHINAGAR",
            "project_name": "HALOL WSS SWAP-III under AMRUT 2.0 & SJMMSVY UGD PHASE II",
            "work_name": "Full work description",
            "work_order": "Work Order No. XYZ / 2024",
            "agency": "Aditi Construction Pvt Ltd",
            "items": [
                {
                    "sr_no": "1",
                    "schedule": "B-1",
                    "description": "DI K7 Pipe 100mm dia laying",
                    "unit": "RM",
                    "tender_qty": 2400,
                    "tender_rate": 485.0,
                    "prev_qty": 1200,
                    "prev_amount": 582000.0,
                    "this_qty": 650,
                    "this_amount": 315250.0,
                    "upto_qty": 1850,
                    "upto_amount": 897250.0
                },
                ...
            ]
        }

    Returns:
        bytes: Excel file as bytes ready for download / file write.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    bill_no = bill_data.get("bill_no", "1")
    period_from = bill_data.get("bill_period_from", "")
    period_to = bill_data.get("bill_period_to", "")
    bill_date = bill_data.get("bill_date", "")
    company = bill_data.get("company_name", "GUJARAT URBAN DEVELOPMENT COMPANY, GANDHINAGAR")
    project = bill_data.get("project_name", "HALOL WSS SWAP-III under AMRUT 2.0 & SJMMSVY UGD PHASE II")
    work_name = bill_data.get("work_name", "")
    work_order = bill_data.get("work_order", "")
    agency = bill_data.get("agency", "")
    items = bill_data.get("items", [])

    # ------------------------------------------------------------------ #
    #  Sheet 1 – Payment Abstract                                          #
    # ------------------------------------------------------------------ #
    ws1 = wb.create_sheet("Payment Abstract")

    col_widths_1 = [6, 16, 36, 18, 18, 18, 18]
    for i, w in enumerate(col_widths_1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    _merge(ws1, 1, 1, 7, company, bold=True, size=14, height=24)
    _merge(ws1, 2, 1, 7, project, bold=True, size=11, height=20)
    _merge(ws1, 3, 1, 7, work_name, bold=False, size=10, height=28, wrap=True)
    _merge(ws1, 4, 1, 7, work_order, bold=False, height=18)
    _merge(ws1, 5, 1, 7, f"Agency: {agency}", bold=False, height=18)
    _merge(ws1, 6, 1, 7, f"RA Bill - {bill_no}", bold=True, size=12, height=20)
    _merge(ws1, 7, 1, 7, "GROSS PAYMENT SUMMARY OF ABSTRACT",
           bold=True, size=12, fill=HEADER_FILL, height=20)

    # Row 8 – Table headers
    headers = ["Sr.No", "Schedule No", "Nagarpalika", "BOQ Quoted Amt",
               "Upto Date Amt", "Prev Bill Amt", "This Bill Amt"]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=8, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    ws1.row_dimensions[8].height = 22

    # Group items by schedule
    schedules = {}
    for item in items:
        sched = item.get("schedule", "B-1")
        schedules.setdefault(sched, []).append(item)

    row = 9
    grand_boq = grand_upto = grand_prev = grand_this = 0.0

    for sched, sched_items in schedules.items():
        boq_amt = sum(i.get("tender_qty", 0) * i.get("tender_rate", 0) for i in sched_items)
        upto_amt = sum(i.get("upto_amount", 0) for i in sched_items)
        prev_amt = sum(i.get("prev_amount", 0) for i in sched_items)
        this_amt = sum(i.get("this_amount", 0) for i in sched_items)

        data_row = [row - 8, sched, f"Halol {sched}", boq_amt, upto_amt, prev_amt, this_amt]
        for col, val in enumerate(data_row, 1):
            c = ws1.cell(row=row, column=col, value=val)
            c.border = THIN_BORDER
            if col >= 4:
                c.number_format = AMOUNT_FORMAT
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

        grand_boq += boq_amt
        grand_upto += upto_amt
        grand_prev += prev_amt
        grand_this += this_amt
        row += 1

    # Grand total row
    total_data = ["", "TOTAL", "", grand_boq, grand_upto, grand_prev, grand_this]
    for col, val in enumerate(total_data, 1):
        c = ws1.cell(row=row, column=col, value=val)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        if col >= 4:
            c.number_format = AMOUNT_FORMAT
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws1.freeze_panes = "A9"

    # ------------------------------------------------------------------ #
    #  Sheet 2 – Statement of Accounts                                     #
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet("STATEMENT OF ACCOUNTS")
    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 20

    _merge(ws2, 1, 1, 3, "STATEMENT OF ACCOUNTS", bold=True, size=13,
           fill=HEADER_FILL, height=22)
    _merge(ws2, 2, 1, 3, f"Project: {project}", bold=False, height=18)
    _merge(ws2, 3, 1, 3, f"RA Bill No: {bill_no} | Period: {period_from} to {period_to} | Date: {bill_date}", bold=False, height=18)

    # Nagarpalika-wise totals
    halol_wss = sum(i.get("upto_amount", 0) for i in items if "WSS" in i.get("schedule", "").upper() or i.get("schedule", "").startswith("B-1"))
    halol_ugd = sum(i.get("upto_amount", 0) for i in items if "UGD" in i.get("schedule", "").upper() or i.get("schedule", "").startswith("B-2"))
    total_c = halol_wss + halol_ugd
    tp_deduction = total_c * 0.036
    e_val = total_c - tp_deduction
    price_variation = 0.0
    g_val = e_val + price_variation
    retention = g_val * 0.05
    net_payable = g_val - retention

    soa_rows = [
        ("A", f"Halol WSS Amount (Gross, upto date)",     halol_wss),
        ("B", f"Halol UGD Amount (Gross, upto date)",     halol_ugd),
        ("C", "Total (A + B)",                            total_c),
        ("D", "T.P. -3.60% of C",                       -tp_deduction),
        ("E", "C - D",                                    e_val),
        ("F", "Price Variation (Clause-59)",              price_variation),
        ("G", "E + F",                                    g_val),
        ("H", "5% Retention of G",                       -retention),
        ("I", "Net Payable (G - H)  [Excl. GST]",        net_payable),
    ]

    for r_idx, (label, desc, amt) in enumerate(soa_rows, 4):
        is_net = label == "I"
        ws2.cell(row=r_idx, column=1, value=label).border = THIN_BORDER
        ws2.cell(row=r_idx, column=1).font = Font(bold=True)
        ws2.cell(row=r_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")

        dc = ws2.cell(row=r_idx, column=2, value=desc)
        dc.border = THIN_BORDER
        dc.font = Font(bold=is_net)
        if is_net:
            dc.fill = HEADER_FILL

        ac = ws2.cell(row=r_idx, column=3, value=amt)
        ac.border = THIN_BORDER
        ac.number_format = AMOUNT_FORMAT
        ac.alignment = Alignment(horizontal="right", vertical="center")
        ac.font = Font(bold=is_net)
        if is_net:
            ac.fill = HEADER_FILL

    ws2.freeze_panes = "A4"

    # ------------------------------------------------------------------ #
    #  Sheet 3 – Abstract Sheet                                            #
    # ------------------------------------------------------------------ #
    ws3 = wb.create_sheet("Abstract Sheet")

    col_widths_3 = [6, 40, 8, 12, 12, 12, 14, 12, 14, 16]
    for i, w in enumerate(col_widths_3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    _merge(ws3, 1, 1, 10, "ABSTRACT SHEET – BILL OF QUANTITIES",
           bold=True, size=13, fill=HEADER_FILL, height=22)
    _merge(ws3, 2, 1, 10, project, bold=False, height=18)
    _merge(ws3, 3, 1, 10,
           f"RA Bill No: {bill_no} | Period: {period_from} to {period_to}", bold=False, height=18)

    abs_hdrs = ["Sr.No", "Description", "Unit", "Tender Qty", "Tender Rate",
                "Prev Qty", "Prev Amt", "This Qty", "This Amt", "Upto Date Amt"]
    for col, h in enumerate(abs_hdrs, 1):
        c = ws3.cell(row=4, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
    ws3.row_dimensions[4].height = 22

    # Group by schedule and add sub-totals
    current_row = 5
    grand_upto_abs = 0.0

    for sched, sched_items in schedules.items():
        # Schedule header
        ws3.cell(row=current_row, column=1, value="").border = THIN_BORDER
        sched_hdr = ws3.cell(row=current_row, column=2, value=f"Schedule {sched}")
        sched_hdr.font = Font(bold=True, italic=True)
        sched_hdr.fill = PatternFill(start_color="EEF4FF", end_color="EEF4FF", fill_type="solid")
        sched_hdr.border = THIN_BORDER
        for col in range(3, 11):
            c = ws3.cell(row=current_row, column=col)
            c.border = THIN_BORDER
            c.fill = PatternFill(start_color="EEF4FF", end_color="EEF4FF", fill_type="solid")
        current_row += 1

        sched_upto = 0.0
        for item in sched_items:
            tender_qty = item.get("tender_qty", 0)
            tender_rate = item.get("tender_rate", 0)
            prev_qty = item.get("prev_qty", 0)
            prev_amt = item.get("prev_amount", 0)
            this_qty = item.get("this_qty", 0)
            this_amt = item.get("this_amount", 0)
            upto_qty = item.get("upto_qty", 0)
            upto_amt = item.get("upto_amount", 0)

            row_vals = [item.get("sr_no", current_row - 4), item.get("description", ""),
                        item.get("unit", ""), tender_qty, tender_rate,
                        prev_qty, prev_amt, this_qty, this_amt, upto_amt]

            for col, val in enumerate(row_vals, 1):
                c = ws3.cell(row=current_row, column=col, value=val)
                c.border = THIN_BORDER
                c.alignment = Alignment(
                    horizontal="right" if col >= 4 else ("center" if col == 1 else "left"),
                    vertical="center",
                    wrap_text=(col == 2)
                )
                if col in (5, 7, 9, 10):
                    c.number_format = AMOUNT_FORMAT
                elif col in (4, 6, 8):
                    c.number_format = "#,##0.00"

            sched_upto += upto_amt
            current_row += 1

        # Sub-total row
        st_vals = ["", f"Sub-Total {sched}", "", "", "", "", "", "", "", sched_upto]
        for col, val in enumerate(st_vals, 1):
            c = ws3.cell(row=current_row, column=col, value=val)
            c.font = Font(bold=True)
            c.fill = HEADER_FILL
            c.border = THIN_BORDER
            if col == 10:
                c.number_format = AMOUNT_FORMAT
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
        grand_upto_abs += sched_upto
        current_row += 1

    # Grand total
    gt_vals = ["", "GRAND TOTAL", "", "", "", "", "", "", "", grand_upto_abs]
    for col, val in enumerate(gt_vals, 1):
        c = ws3.cell(row=current_row, column=col, value=val)
        c.font = Font(bold=True, size=12)
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        if col == 10:
            c.number_format = AMOUNT_FORMAT
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws3.freeze_panes = "A5"

    # ------------------------------------------------------------------ #
    #  Output                                                              #
    # ------------------------------------------------------------------ #
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ------------------------------------------------------------------ #
#  Standalone test / demo                                              #
# ------------------------------------------------------------------ #
if __name__ == "__main__":
    sample_data = {
        "bill_no": "3",
        "bill_period_from": "2024-10-01",
        "bill_period_to": "2025-03-31",
        "bill_date": "2025-04-01",
        "company_name": "GUJARAT URBAN DEVELOPMENT COMPANY, GANDHINAGAR",
        "project_name": "HALOL WSS SWAP-III under AMRUT 2.0 & SJMMSVY UGD PHASE II",
        "work_name": "Construction of Water Supply Scheme for Halol Nagarpalika including DI pipes, valves, chambers and appurtenances",
        "work_order": "GUDCL/AMRUT/HALOL/WSS/WO-2024/001 Dated: 15-Apr-2024",
        "agency": "Aditi Construction Pvt Ltd, Ahmedabad",
        "items": [
            {
                "sr_no": "1", "schedule": "B-1",
                "description": "DI K7 Pipe 100mm dia laying including jointing",
                "unit": "RM",
                "tender_qty": 2400, "tender_rate": 485.0,
                "prev_qty": 1200, "prev_amount": 582000.0,
                "this_qty": 650, "this_amount": 315250.0,
                "upto_qty": 1850, "upto_amount": 897250.0
            },
            {
                "sr_no": "2", "schedule": "B-1",
                "description": "DI K7 Pipe 150mm dia laying including jointing",
                "unit": "RM",
                "tender_qty": 1200, "tender_rate": 720.0,
                "prev_qty": 600, "prev_amount": 432000.0,
                "this_qty": 300, "this_amount": 216000.0,
                "upto_qty": 900, "upto_amount": 648000.0
            },
            {
                "sr_no": "3", "schedule": "B-1",
                "description": "Gate Valve 100mm dia (IS:14846) installation",
                "unit": "Nos",
                "tender_qty": 24, "tender_rate": 12500.0,
                "prev_qty": 12, "prev_amount": 150000.0,
                "this_qty": 6, "this_amount": 75000.0,
                "upto_qty": 18, "upto_amount": 225000.0
            },
            {
                "sr_no": "4", "schedule": "B-2",
                "description": "Earthwork excavation in soft soil for UGD",
                "unit": "Cum",
                "tender_qty": 5000, "tender_rate": 185.0,
                "prev_qty": 2500, "prev_amount": 462500.0,
                "this_qty": 1200, "this_amount": 222000.0,
                "upto_qty": 3700, "upto_amount": 684500.0
            },
            {
                "sr_no": "5", "schedule": "B-2",
                "description": "RCC M20 grade concrete for valve chambers",
                "unit": "Cum",
                "tender_qty": 150, "tender_rate": 6800.0,
                "prev_qty": 75, "prev_amount": 510000.0,
                "this_qty": 40, "this_amount": 272000.0,
                "upto_qty": 115, "upto_amount": 782000.0
            },
        ]
    }

    excel_bytes = generate_ra_bill(sample_data)
    output_path = "/tmp/sample_ra_bill.xlsx"
    with open(output_path, "wb") as f:
        f.write(excel_bytes)
    print(f"RA Bill generated: {output_path} ({len(excel_bytes):,} bytes)")
