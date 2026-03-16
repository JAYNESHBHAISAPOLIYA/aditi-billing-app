"""
Government RA Bill Excel Generator using openpyxl
Generates GUDCL-format RA Bill Excel file with 3 sheets:
  1. Payment Abstract
  2. STATEMENT OF ACCOUNTS
  3. Abstract Sheet

Usage:
    from ra_bill_generator import generate_ra_bill
    excel_bytes = generate_ra_bill(bill_data)

    # Or run standalone:
    python ra_bill_generator.py
"""

import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Style helpers ────────────────────────────────────────────────────────────

LIGHT_BLUE_FILL = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

BOLD = Font(bold=True)
BOLD_14 = Font(bold=True, size=14)
BOLD_GREEN = Font(bold=True, color="006600")

AMOUNT_FORMAT = "##,##,##0.00"


def _fmt_inr(value: float) -> str:
    """Format number with Indian comma system."""
    value = float(value or 0)
    is_negative = value < 0
    value = abs(value)
    s = f"{value:.2f}"
    int_part, dec_part = s.split(".")
    # Indian grouping: last 3 then groups of 2
    if len(int_part) > 3:
        result = int_part[-3:]
        int_part = int_part[:-3]
        while int_part:
            result = int_part[-2:] + "," + result
            int_part = int_part[:-2]
    else:
        result = int_part
    return ("-₹" if is_negative else "₹") + result + "." + dec_part


def _apply_header_style(cell, size=11):
    cell.font = Font(bold=True, size=size)
    cell.fill = LIGHT_BLUE_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER


def _apply_data_cell(cell, align=None):
    cell.border = THIN_BORDER
    if align:
        cell.alignment = align


def _merge_and_set(ws, cell_range, value, font=None, fill=None, alignment=None, border=None):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_payment_abstract(wb: Workbook, bill_data: dict) -> None:
    """Sheet 1 – Payment Abstract."""
    ws = wb.create_sheet("Payment Abstract")

    company_name = bill_data.get("company_name", "GUJARAT URBAN DEVELOPMENT COMPANY, GANDHINAGAR")
    project_title = bill_data.get("project_title", "HALOL WSS SWAP-III under AMRUT 2.0 & SJMMSVY UGD PHASE II")
    work_name = bill_data.get("work_name", "Construction of Water Supply Distribution System")
    work_order = bill_data.get("work_order", "N/A")
    agency = bill_data.get("agency", "As per Contract")
    bill_no = bill_data.get("bill_no", 1)
    items = bill_data.get("items", [])

    # Column widths
    col_widths = [6, 14, 35, 16, 16, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1 – Company name
    _merge_and_set(ws, "A1:G1", company_name, font=BOLD_14, fill=LIGHT_BLUE_FILL, alignment=CENTER_ALIGN)

    # Row 2 – Project title
    _merge_and_set(ws, "A2:G2", project_title, font=BOLD, fill=LIGHT_BLUE_FILL, alignment=CENTER_ALIGN)

    # Row 3 – Full work name
    _merge_and_set(ws, "A3:G3", work_name, alignment=Alignment(wrap_text=True))

    # Row 4 – Work order details
    _merge_and_set(ws, "A4:G4", f"Work Order No: {work_order}")

    # Row 5 – Agency
    _merge_and_set(ws, "A5:G5", f"Agency/Contractor: {agency}")

    # Row 6 – Bill number
    _merge_and_set(ws, "A6:G6", f"RA Bill - {bill_no}", font=BOLD, alignment=CENTER_ALIGN)

    # Row 7 – Summary title
    _merge_and_set(ws, "A7:G7", "GROSS PAYMENT SUMMARY OF ABSTRACT",
                   font=BOLD, fill=LIGHT_BLUE_FILL, alignment=CENTER_ALIGN)

    # Row 8 – Table header
    headers = ["Sr.No", "Schedule No", "Nagarpalika", "BOQ Quoted Amt",
               "Upto Date Amt", "Prev Bill Amt", "This Bill Amt"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=h)
        _apply_header_style(cell)

    # Freeze panes at row 8
    ws.freeze_panes = "A9"

    boq_quoted = sum(i.get("tender_qty", 0) * i.get("tender_rate", 0) for i in items)
    upto_total = sum(i.get("upto_amount", 0) for i in items)
    prev_total = sum(i.get("prev_amount", 0) for i in items)
    this_total = sum(i.get("this_amount", 0) for i in items)

    # Row 9 – WSS data
    row9 = [1, "B-1", bill_data.get("nagarpalika_wss", "Halol WSS"),
            _fmt_inr(boq_quoted), _fmt_inr(upto_total), _fmt_inr(prev_total), _fmt_inr(this_total)]
    for col, val in enumerate(row9, 1):
        cell = ws.cell(row=9, column=col, value=val)
        _apply_data_cell(cell, RIGHT_ALIGN if col >= 4 else CENTER_ALIGN)

    # Row 10 – UGD data (placeholder)
    row10 = [2, "B-2", bill_data.get("nagarpalika_ugd", "Halol UGD"),
             _fmt_inr(0), _fmt_inr(0), _fmt_inr(0), _fmt_inr(0)]
    for col, val in enumerate(row10, 1):
        cell = ws.cell(row=10, column=col, value=val)
        _apply_data_cell(cell, RIGHT_ALIGN if col >= 4 else CENTER_ALIGN)


def _build_statement_of_accounts(wb: Workbook, bill_data: dict) -> None:
    """Sheet 2 – Statement of Accounts."""
    ws = wb.create_sheet("STATEMENT OF ACCOUNTS")

    col_widths = [5, 55, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    items = bill_data.get("items", [])
    bill_no = bill_data.get("bill_no", 1)
    period_from = bill_data.get("bill_period_from", "-")
    period_to = bill_data.get("bill_period_to", "-")
    nagarpalika_wss = bill_data.get("nagarpalika_wss", "Halol WSS")
    nagarpalika_ugd = bill_data.get("nagarpalika_ugd", "Halol UGD")

    this_total = sum(i.get("this_amount", 0) for i in items)

    # Title
    _merge_and_set(ws, "B1:C1", "STATEMENT OF ACCOUNTS",
                   font=Font(bold=True, size=12), fill=LIGHT_BLUE_FILL, alignment=CENTER_ALIGN)

    ws.cell(row=2, column=2,
            value=f"RA Bill No: {bill_no}  |  Period: {period_from} to {period_to}")
    ws.row_dimensions[3].height = 6  # spacer

    # Amounts A–I
    gross_a = this_total
    gross_b = bill_data.get("ugd_amount", 0)
    total_c = gross_a + gross_b
    tp_d = total_c * 0.036
    e = total_c - tp_d
    pv_f = bill_data.get("price_variation", 0)
    g = e + pv_f
    retention_h = g * 0.05
    net_i = g - retention_h

    rows = [
        ("A", f"{nagarpalika_wss} amount", gross_a),
        ("B", f"{nagarpalika_ugd} amount", gross_b),
        ("C", "Total (A+B)", total_c),
        ("D", "T.P. -3.60% of C", -tp_d),
        ("E", "C - D", e),
        ("F", "Price Variation (Clause-59)", pv_f),
        ("G", "E + F", g),
        ("H", "5% Retention of G", -retention_h),
        ("I", "G - H  ← Net Payable (Excl. GST)", net_i),
    ]

    start_row = 4
    for idx, (label, desc, amt) in enumerate(rows):
        r = start_row + idx
        lc = ws.cell(row=r, column=1, value=label)
        dc = ws.cell(row=r, column=2, value=desc)
        ac = ws.cell(row=r, column=3, value=_fmt_inr(amt))
        lc.font = BOLD
        ac.alignment = RIGHT_ALIGN
        for c in [lc, dc, ac]:
            c.border = THIN_BORDER
        if label == "I":
            for c in [lc, dc, ac]:
                c.font = BOLD_GREEN
                c.fill = GREEN_FILL


def _build_abstract_sheet(wb: Workbook, bill_data: dict) -> None:
    """Sheet 3 – Abstract Sheet with all BOQ items."""
    ws = wb.create_sheet("Abstract Sheet")

    col_widths = [6, 12, 40, 8, 12, 12, 14, 12, 14, 12, 14, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    company_name = bill_data.get("company_name", "GUJARAT URBAN DEVELOPMENT COMPANY, GANDHINAGAR")
    project_title = bill_data.get("project_title", "HALOL WSS SWAP-III under AMRUT 2.0 & SJMMSVY UGD PHASE II")
    bill_no = bill_data.get("bill_no", 1)
    period_from = bill_data.get("bill_period_from", "-")
    period_to = bill_data.get("bill_period_to", "-")
    bill_date = bill_data.get("bill_date", "-")
    items = bill_data.get("items", [])

    # Rows 1–3: title
    _merge_and_set(ws, "A1:M1", company_name, font=Font(bold=True, size=13),
                   fill=LIGHT_BLUE_FILL, alignment=CENTER_ALIGN)
    _merge_and_set(ws, "A2:M2", project_title, font=BOLD, alignment=CENTER_ALIGN)
    _merge_and_set(ws, "A3:M3",
                   f"RA Bill - {bill_no}  |  Period: {period_from} to {period_to}  |  Date: {bill_date}")

    # Row 4: blank spacer
    ws.row_dimensions[4].height = 6

    # Row 5: column headers
    headers = [
        "Sr", "Schedule", "Description", "Unit",
        "Tender Qty", "Tender Rate", "Tender Amt",
        "Prev Qty", "Prev Amt",
        "This Qty", "This Amt",
        "Upto Qty", "Upto Amt",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        _apply_header_style(cell)

    # Freeze panes
    ws.freeze_panes = "A6"

    # Group items by schedule prefix
    from collections import defaultdict
    schedule_groups = defaultdict(list)
    for item in items:
        schedule = item.get("schedule", "")
        prefix = schedule.split("-")[0] if "-" in schedule else schedule or "General"
        schedule_groups[prefix].append(item)

    current_row = 6
    grand_totals = {k: 0 for k in ["tender_qty", "tender_amt", "prev_qty", "prev_amt",
                                    "this_qty", "this_amt", "upto_qty", "upto_amt"]}

    for schedule_key, group_items in sorted(schedule_groups.items()):
        sub_totals = {k: 0 for k in grand_totals}

        for item in group_items:
            tender_amt = item.get("tender_qty", 0) * item.get("tender_rate", 0)
            row_vals = [
                item.get("sr_no", ""),
                item.get("schedule", ""),
                item.get("description", ""),
                item.get("unit", ""),
                item.get("tender_qty", 0),
                item.get("tender_rate", 0),
                tender_amt,
                item.get("prev_qty", 0),
                item.get("prev_amount", 0),
                item.get("this_qty", 0),
                item.get("this_amount", 0),
                item.get("upto_qty", 0),
                item.get("upto_amount", 0),
            ]
            for col, val in enumerate(row_vals, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.border = THIN_BORDER
                if col >= 5:
                    cell.alignment = RIGHT_ALIGN
                    if col in [7, 9, 11, 13]:
                        cell.number_format = AMOUNT_FORMAT
                else:
                    cell.alignment = LEFT_ALIGN if col == 3 else CENTER_ALIGN

            sub_totals["tender_qty"] += item.get("tender_qty", 0)
            sub_totals["tender_amt"] += tender_amt
            sub_totals["prev_qty"] += item.get("prev_qty", 0)
            sub_totals["prev_amt"] += item.get("prev_amount", 0)
            sub_totals["this_qty"] += item.get("this_qty", 0)
            sub_totals["this_amt"] += item.get("this_amount", 0)
            sub_totals["upto_qty"] += item.get("upto_qty", 0)
            sub_totals["upto_amt"] += item.get("upto_amount", 0)
            current_row += 1

        # Sub-total row for schedule group
        if len(schedule_groups) > 1:
            sub_row = [
                "", "", f"Sub-Total {schedule_key}", "",
                sub_totals["tender_qty"], "", sub_totals["tender_amt"],
                sub_totals["prev_qty"], sub_totals["prev_amt"],
                sub_totals["this_qty"], sub_totals["this_amt"],
                sub_totals["upto_qty"], sub_totals["upto_amt"],
            ]
            for col, val in enumerate(sub_row, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.font = BOLD
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = THIN_BORDER
                if col >= 5:
                    cell.alignment = RIGHT_ALIGN
                    if col in [7, 9, 11, 13]:
                        cell.number_format = AMOUNT_FORMAT
            current_row += 1

        for k in grand_totals:
            grand_totals[k] += sub_totals[k]

    # Grand total row
    gt_row = [
        "", "", "GRAND TOTAL", "",
        grand_totals["tender_qty"], "", grand_totals["tender_amt"],
        grand_totals["prev_qty"], grand_totals["prev_amt"],
        grand_totals["this_qty"], grand_totals["this_amt"],
        grand_totals["upto_qty"], grand_totals["upto_amt"],
    ]
    for col, val in enumerate(gt_row, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = BOLD
        cell.fill = LIGHT_BLUE_FILL
        cell.border = THIN_BORDER
        if col >= 5:
            cell.alignment = RIGHT_ALIGN
            if col in [7, 9, 11, 13]:
                cell.number_format = AMOUNT_FORMAT


# ── Main public API ────────────────────────────────────────────────────────────

def generate_ra_bill(bill_data: dict) -> bytes:
    """
    Generate a Government RA Bill Excel file and return as bytes.

    bill_data dict keys:
        company_name        str  – Header row 1
        project_title       str  – Header row 2
        work_name           str  – Full work description
        work_order          str  – Work order number
        agency              str  – Contractor/Agency name
        bill_no             int  – RA Bill number
        bill_period_from    str  – Period start (YYYY-MM-DD)
        bill_period_to      str  – Period end (YYYY-MM-DD)
        bill_date           str  – Bill date (YYYY-MM-DD)
        nagarpalika_wss     str  – WSS Nagarpalika name
        nagarpalika_ugd     str  – UGD Nagarpalika name
        ugd_amount          float – UGD gross amount
        price_variation     float – Price variation (Clause-59)
        items               list  – BOQ items:
            sr_no, schedule, description, unit,
            tender_qty, tender_rate,
            prev_qty, prev_amount,
            this_qty, this_amount,
            upto_qty, upto_amount

    Returns:
        bytes: Excel file content
    """
    wb = Workbook()
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)

    _build_payment_abstract(wb, bill_data)
    _build_statement_of_accounts(wb, bill_data)
    _build_abstract_sheet(wb, bill_data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Standalone test ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    sample_data = {
        "company_name": "GUJARAT URBAN DEVELOPMENT COMPANY, GANDHINAGAR",
        "project_title": "HALOL WSS SWAP-III under AMRUT 2.0 & SJMMSVY UGD PHASE II",
        "work_name": "Construction of Water Supply Distribution System for Halol Nagarpalika",
        "work_order": "GUDCL/HALOL/WSS/2024/001",
        "agency": "ABC Construction Pvt. Ltd.",
        "bill_no": 3,
        "bill_period_from": "2024-01-01",
        "bill_period_to": "2024-01-31",
        "bill_date": "2024-02-05",
        "nagarpalika_wss": "Halol WSS",
        "nagarpalika_ugd": "Halol UGD",
        "ugd_amount": 0,
        "price_variation": 0,
        "items": [
            {"sr_no": 1, "schedule": "B-1", "description": "DI K7 pipe 100mm dia laying", "unit": "RM", "tender_qty": 2400, "tender_rate": 485, "prev_qty": 1200, "prev_amount": 582000, "this_qty": 650, "this_amount": 315250, "upto_qty": 1850, "upto_amount": 897250},
            {"sr_no": 2, "schedule": "B-1", "description": "DI K9 pipe 150mm dia laying", "unit": "RM", "tender_qty": 800, "tender_rate": 720, "prev_qty": 400, "prev_amount": 288000, "this_qty": 200, "this_amount": 144000, "upto_qty": 600, "upto_amount": 432000},
            {"sr_no": 3, "schedule": "B-2", "description": "Sluice valve 100mm flanged", "unit": "Nos", "tender_qty": 12, "tender_rate": 8500, "prev_qty": 6, "prev_amount": 51000, "this_qty": 3, "this_amount": 25500, "upto_qty": 9, "upto_amount": 76500},
            {"sr_no": 4, "schedule": "B-2", "description": "Brick masonry valve chamber", "unit": "Nos", "tender_qty": 8, "tender_rate": 12500, "prev_qty": 4, "prev_amount": 50000, "this_qty": 2, "this_amount": 25000, "upto_qty": 6, "upto_amount": 75000},
            {"sr_no": 5, "schedule": "R-1", "description": "Earthwork excavation ordinary soil", "unit": "Cum", "tender_qty": 3600, "tender_rate": 185, "prev_qty": 1800, "prev_amount": 333000, "this_qty": 975, "this_amount": 180375, "upto_qty": 2775, "upto_amount": 513375},
            {"sr_no": 6, "schedule": "R-1", "description": "Refilling excavated earth with compaction", "unit": "Cum", "tender_qty": 3200, "tender_rate": 95, "prev_qty": 1600, "prev_amount": 152000, "this_qty": 850, "this_amount": 80750, "upto_qty": 2450, "upto_amount": 232750},
            {"sr_no": 7, "schedule": "P-1", "description": "UPVC pressure pipe 110mm Class C", "unit": "RM", "tender_qty": 1500, "tender_rate": 285, "prev_qty": 750, "prev_amount": 213750, "this_qty": 400, "this_amount": 114000, "upto_qty": 1150, "upto_amount": 327750},
            {"sr_no": 8, "schedule": "P-1", "description": "HDPE pipe 63mm PN10 house service", "unit": "RM", "tender_qty": 600, "tender_rate": 165, "prev_qty": 0, "prev_amount": 0, "this_qty": 200, "this_amount": 33000, "upto_qty": 200, "upto_amount": 33000},
        ],
    }

    output_bytes = generate_ra_bill(sample_data)
    import tempfile
    import os
    output_path = os.path.join(tempfile.gettempdir(), "RA_Bill_Sample.xlsx")
    with open(output_path, "wb") as f:
        f.write(output_bytes)
    print(f"RA Bill Excel generated: {output_path} ({len(output_bytes):,} bytes)")
