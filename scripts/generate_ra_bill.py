"""
Government RA Bill Excel Generator
Generates RA Bill in GUDCL/AMRUT 2.0 format using openpyxl.

Usage:
    from scripts.generate_ra_bill import generate_ra_bill
    excel_bytes = generate_ra_bill(bill_data)

    # Or run standalone:
    python scripts/generate_ra_bill.py
"""

from io import BytesIO
from typing import Any
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ─── Style constants ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
SUBTOTAL_FILL = PatternFill(start_color="DDEECC", end_color="DDEECC", fill_type="solid")
GRANDTOTAL_FILL = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
SCHEDULE_FILL = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")

THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BOLD = Font(bold=True)
BOLD_LG = Font(bold=True, size=14)
BOLD_MD = Font(bold=True, size=12)
BOLD_SM = Font(bold=True, size=11)

CENTER = Alignment(horizontal="center", vertical="middle", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="middle", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="middle")

INR_FMT = '##,##,##0.00'


def _cell(ws, row: int, col: int, value: Any = None, *,
          font=None, fill=None, align=None, border=None, num_fmt=None):
    """Set a single cell's value and formatting."""
    c = ws.cell(row=row, column=col, value=value)
    if font:    c.font = font
    if fill:    c.fill = fill
    if align:   c.alignment = align
    if border:  c.border = border
    if num_fmt: c.number_format = num_fmt
    return c


def _merge(ws, row: int, col1: int, col2: int, value: Any = None, *,
           font=None, fill=None, align=CENTER, border=None):
    """Merge cells and format."""
    ws.merge_cells(start_row=row, start_column=col1,
                   end_row=row, end_column=col2)
    return _cell(ws, row, col1, value, font=font, fill=fill, align=align, border=border)


def _border_row(ws, row: int, col1: int, col2: int):
    """Apply thin border to a range of cells in a row."""
    for c in range(col1, col2 + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER


# ─── Sheet 1: Payment Abstract ────────────────────────────────────────────────

def _build_payment_abstract(wb: Workbook, bill_data: dict):
    ws = wb.create_sheet("Payment Abstract")

    # Column widths: A-J
    widths = [8, 15, 30, 18, 18, 18, 18, 12, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    items = bill_data.get("items", [])
    boq_quoted = sum(i.get("tender_qty", 0) * i.get("tender_rate", 0) for i in items)
    upto_date  = sum(i.get("upto_amount", 0) for i in items)
    prev_amt   = sum(i.get("prev_amount", 0) for i in items)
    this_amt   = sum(i.get("this_amount", 0) for i in items)

    # Row 1 – Company name
    _merge(ws, 1, 1, 10,
           "GUJARAT URBAN DEVELOPMENT COMPANY, GANDHINAGAR",
           font=BOLD_LG, fill=HEADER_FILL)
    ws.row_dimensions[1].height = 24

    # Row 2 – Project sub-title
    _merge(ws, 2, 1, 10,
           bill_data.get("project_subtitle",
                         "HALOL WSS SWAP-III under AMRUT 2.0 & SJMMSVY UGD PHASE II"),
           font=BOLD_SM)
    ws.row_dimensions[2].height = 20

    # Row 3 – Full work name
    _merge(ws, 3, 1, 10,
           f"Work: {bill_data.get('site_name', '')}  |  "
           f"Location: {bill_data.get('site_location', '')}",
           align=Alignment(horizontal="left", vertical="middle", wrap_text=True))
    ws.row_dimensions[3].height = 30

    # Row 4 – Work order details
    _merge(ws, 4, 1, 10,
           f"Tender No: {bill_data.get('tender_number', 'N/A')}  |  "
           f"Bill Period: {bill_data.get('bill_period_from', '')} to "
           f"{bill_data.get('bill_period_to', '')}  |  Date: {bill_data.get('bill_date', '')}",
           align=LEFT)

    # Row 5 – Agency
    _merge(ws, 5, 1, 10,
           f"Agency / Contractor: {bill_data.get('contractor_name', '')}",
           align=LEFT)

    # Row 6 – RA Bill number
    _merge(ws, 6, 1, 10,
           f"RA Bill - {bill_data.get('bill_no', '1')}",
           font=BOLD)

    # Row 7 – Section header
    _merge(ws, 7, 1, 10,
           "GROSS PAYMENT SUMMARY OF ABSTRACT",
           font=BOLD, fill=HEADER_FILL, border=THIN_BORDER)
    ws.row_dimensions[7].height = 22

    # Row 8 – Column headers
    headers = ["Sr.No", "Schedule No", "Work / Nagarpalika",
               "BOQ Quoted Amt (₹)", "Upto Date Amt (₹)",
               "Prev Bill Amt (₹)", "This Bill Amt (₹)"]
    for col, h in enumerate(headers, start=1):
        _cell(ws, 8, col, h, font=BOLD, fill=HEADER_FILL,
              align=CENTER, border=THIN_BORDER)
    ws.row_dimensions[8].height = 30

    # Row 9 – Site data
    row9_data = [1, bill_data.get("site_name", "")[:20],
                 bill_data.get("site_name", ""),
                 boq_quoted, upto_date, prev_amt, this_amt]
    for col, val in enumerate(row9_data, start=1):
        fmt = INR_FMT if col >= 4 else None
        alg = RIGHT if col >= 4 else LEFT
        _cell(ws, 9, col, val, border=THIN_BORDER, align=alg, num_fmt=fmt)
    ws.row_dimensions[9].height = 20

    # Freeze panes below row 8
    ws.freeze_panes = "A9"


# ─── Sheet 2: Statement of Accounts ──────────────────────────────────────────

def _build_statement_of_accounts(wb: Workbook, bill_data: dict):
    ws = wb.create_sheet("STATEMENT OF ACCOUNTS")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 22

    items    = bill_data.get("items", [])
    upto_amt = sum(i.get("upto_amount", 0) for i in items)

    tp_rate   = 0.036          # 3.60%
    tp        = upto_amt * tp_rate
    after_tp  = upto_amt - tp
    pv        = 0.0            # Price Variation – user can override
    after_pv  = after_tp + pv
    retention = after_pv * 0.05
    net_pay   = after_pv - retention

    _merge(ws, 1, 1, 3, "STATEMENT OF ACCOUNTS",
           font=BOLD_MD, fill=HEADER_FILL)
    ws.row_dimensions[1].height = 26

    _merge(ws, 2, 1, 3,
           f"RA Bill No: {bill_data.get('bill_no', '1')}  |  Date: {bill_data.get('bill_date', '')}",
           align=CENTER)
    ws.row_dimensions[2].height = 18

    for col, h in enumerate(["Row", "Description", "Amount (₹)"], start=1):
        _cell(ws, 3, col, h, font=BOLD, fill=HEADER_FILL,
              align=CENTER, border=THIN_BORDER)
    ws.row_dimensions[3].height = 22

    soa_rows = [
        ("A", f"{bill_data.get('site_name', 'Site')} amount", upto_amt),
        ("B", "Other works amount",                            0.0),
        ("C", "Total (A + B)",                                upto_amt),
        ("D", f"T.P. -{tp_rate*100:.2f}% of C",              -tp),
        ("E", "C − D",                                        after_tp),
        ("F", "Price Variation (Clause-59)",                  pv),
        ("G", "E + F",                                        after_pv),
        ("H", "5% Retention of G",                           -retention),
        ("I", "G − H  ← Net Payable (excl. GST)",            net_pay),
    ]

    for idx, (label, desc, amt) in enumerate(soa_rows):
        r = 4 + idx
        is_net = label == "I"
        fnt  = BOLD if is_net else None
        fill = HIGHLIGHT_FILL if is_net else None
        _cell(ws, r, 1, label, font=fnt, fill=fill, border=THIN_BORDER, align=CENTER)
        _cell(ws, r, 2, desc,  font=fnt, fill=fill, border=THIN_BORDER, align=LEFT)
        _cell(ws, r, 3, amt,   font=fnt, fill=fill, border=THIN_BORDER,
              align=RIGHT, num_fmt=INR_FMT)
        ws.row_dimensions[r].height = 20


# ─── Sheet 3: Abstract Sheet (BOQ items) ─────────────────────────────────────

def _build_abstract_sheet(wb: Workbook, bill_data: dict):
    ws = wb.create_sheet("Abstract Sheet")

    col_widths = [8, 12, 40, 10, 12, 14, 14, 12, 14, 12, 14, 12, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _merge(ws, 1, 1, 13,
           "ABSTRACT SHEET – BILL OF QUANTITIES",
           font=BOLD_MD, fill=HEADER_FILL)
    ws.row_dimensions[1].height = 26

    _merge(ws, 2, 1, 13,
           f"Site: {bill_data.get('site_name', '')}  |  "
           f"RA Bill No: {bill_data.get('bill_no', '1')}  |  Date: {bill_data.get('bill_date', '')}",
           align=CENTER)
    ws.row_dimensions[2].height = 18

    col_headers = [
        "Sr.No", "Schedule", "Description", "Unit",
        "Tender Qty", "Tender Rate (₹)", "Tender Amt (₹)",
        "Prev Qty", "Prev Amt (₹)",
        "This Qty", "This Amt (₹)",
        "Upto Qty", "Upto Amt (₹)",
    ]
    for col, h in enumerate(col_headers, start=1):
        _cell(ws, 3, col, h, font=BOLD, fill=HEADER_FILL,
              align=CENTER, border=THIN_BORDER)
    ws.row_dimensions[3].height = 30

    # Group items by schedule
    groups: dict[str, list] = {}
    for item in bill_data.get("items", []):
        sch = item.get("schedule", "A")
        groups.setdefault(sch, []).append(item)

    current_row = 4
    g_tender = g_prev = g_this = g_upto = 0.0

    for schedule, items in groups.items():
        # Schedule header
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=13)
        _cell(ws, current_row, 1, f"Schedule: {schedule}",
              font=Font(bold=True, italic=True), fill=SCHEDULE_FILL, align=LEFT)
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        sub_tender = sub_prev = sub_this = sub_upto = 0.0

        for item in items:
            tender_amt = item.get("tender_qty", 0) * item.get("tender_rate", 0)
            row_data = [
                item.get("sr_no", ""),
                item.get("schedule", ""),
                item.get("description", ""),
                item.get("unit", ""),
                item.get("tender_qty", 0),
                item.get("tender_rate", 0),
                tender_amt,
                item.get("prev_qty", 0),
                item.get("prev_amount", 0),
                item.get("this_qty", 0),
                item.get("this_amount", 0),
                item.get("upto_qty", 0),
                item.get("upto_amount", 0),
            ]
            for col, val in enumerate(row_data, start=1):
                fmt = INR_FMT if col in (6, 7, 9, 11, 13) else (
                      "##,##,##0.##" if col in (5, 8, 10, 12) else None)
                alg = RIGHT if col >= 5 else LEFT
                _cell(ws, current_row, col, val,
                      border=THIN_BORDER, align=alg, num_fmt=fmt)
            ws.row_dimensions[current_row].height = 18
            current_row += 1

            sub_tender += tender_amt
            sub_prev   += item.get("prev_amount", 0)
            sub_this   += item.get("this_amount", 0)
            sub_upto   += item.get("upto_amount", 0)

        # Sub-total row
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=6)
        _cell(ws, current_row, 1, f"Sub-Total ({schedule})",
              font=BOLD, fill=SUBTOTAL_FILL, align=LEFT, border=THIN_BORDER)
        for col, val in [(7, sub_tender), (9, sub_prev), (11, sub_this), (13, sub_upto)]:
            _cell(ws, current_row, col, val,
                  font=BOLD, fill=SUBTOTAL_FILL, border=THIN_BORDER,
                  align=RIGHT, num_fmt=INR_FMT)
        _border_row(ws, current_row, 1, 13)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        g_tender += sub_tender
        g_prev   += sub_prev
        g_this   += sub_this
        g_upto   += sub_upto

    # Grand total row
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=6)
    _cell(ws, current_row, 1, "GRAND TOTAL",
          font=BOLD_MD, fill=GRANDTOTAL_FILL, align=LEFT, border=THIN_BORDER)
    for col, val in [(7, g_tender), (9, g_prev), (11, g_this), (13, g_upto)]:
        _cell(ws, current_row, col, val,
              font=BOLD_MD, fill=GRANDTOTAL_FILL, border=THIN_BORDER,
              align=RIGHT, num_fmt=INR_FMT)
    _border_row(ws, current_row, 1, 13)
    ws.row_dimensions[current_row].height = 24

    ws.freeze_panes = "A4"


# ─── Public API ───────────────────────────────────────────────────────────────

def generate_ra_bill(bill_data: dict) -> bytes:
    """
    Generate a Government RA Bill Excel file.

    Parameters
    ----------
    bill_data : dict
        {
          bill_no          : str,
          bill_period_from : str,   # YYYY-MM-DD
          bill_period_to   : str,
          bill_date        : str,
          site_name        : str,
          site_location    : str,
          tender_number    : str,
          contractor_name  : str,
          project_subtitle : str,   # optional
          items : [
            {
              sr_no        : str,
              schedule     : str,   # e.g. "B-1"
              description  : str,
              unit         : str,
              tender_qty   : float,
              tender_rate  : float,
              prev_qty     : float,
              prev_amount  : float,
              this_qty     : float,
              this_amount  : float,
              upto_qty     : float,
              upto_amount  : float,
            }, ...
          ]
        }

    Returns
    -------
    bytes
        Excel file contents.
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _build_payment_abstract(wb, bill_data)
    _build_statement_of_accounts(wb, bill_data)
    _build_abstract_sheet(wb, bill_data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Standalone demo ──────────────────────────────────────────────────────────

if __name__ == "__main__":
    sample_data = {
        "bill_no": "1",
        "bill_period_from": "2024-04-01",
        "bill_period_to": "2024-09-30",
        "bill_date": "2024-10-05",
        "site_name": "Halol WSS SWAP-III",
        "site_location": "Halol, Panchmahal, Gujarat",
        "tender_number": "GJ-2024-WS-HALOL-001",
        "contractor_name": "Aditi Construction Pvt Ltd",
        "project_subtitle": "HALOL WSS SWAP-III under AMRUT 2.0 & SJMMSVY UGD PHASE II",
        "items": [
            {
                "sr_no": "1", "schedule": "B-1 Rising Main",
                "description": "DI K7 Class Pipe 100mm dia (Supply & Laying)",
                "unit": "RM", "tender_qty": 2400, "tender_rate": 485,
                "prev_qty": 1200, "prev_amount": 582000,
                "this_qty": 650, "this_amount": 315250,
                "upto_qty": 1850, "upto_amount": 897250,
            },
            {
                "sr_no": "2", "schedule": "B-1 Rising Main",
                "description": "DI K9 Class Pipe 150mm dia (Supply & Laying)",
                "unit": "RM", "tender_qty": 1200, "tender_rate": 720,
                "prev_qty": 400, "prev_amount": 288000,
                "this_qty": 200, "this_amount": 144000,
                "upto_qty": 600, "upto_amount": 432000,
            },
            {
                "sr_no": "3", "schedule": "B-2 Valves & Fittings",
                "description": "Sluice Valve 100mm dia (Supply & Fixing)",
                "unit": "Nos", "tender_qty": 12, "tender_rate": 8500,
                "prev_qty": 4, "prev_amount": 34000,
                "this_qty": 3, "this_amount": 25500,
                "upto_qty": 7, "upto_amount": 59500,
            },
            {
                "sr_no": "4", "schedule": "B-3 Civil Works",
                "description": "PCC M15 grade (1:2:4) providing & laying",
                "unit": "Cum", "tender_qty": 50, "tender_rate": 4800,
                "prev_qty": 15, "prev_amount": 72000,
                "this_qty": 10, "this_amount": 48000,
                "upto_qty": 25, "upto_amount": 120000,
            },
        ],
    }

    excel_bytes = generate_ra_bill(sample_data)
    out_path = "/tmp/sample_ra_bill.xlsx"
    with open(out_path, "wb") as f:
        f.write(excel_bytes)
    print(f"RA Bill generated: {out_path} ({len(excel_bytes):,} bytes)")
